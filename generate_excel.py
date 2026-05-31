import pandas as pd
from openpyxl import load_workbook
import os
import pymysql


db_config = {
    "host": os.getenv("DB_HOST"),
    "user": os.getenv("DB_USER"),
    "password": os.getenv("DB_PASSWORD"),
    "database": os.getenv("DB_NAME"),
    "cursorclass": pymysql.cursors.DictCursor
}
def get_db_connection():
    return pymysql.connect(
        **db_config,
        connect_timeout=5
    )

def fetch_data(process_id, batch_size=1000):
    conn=None
    cursor= None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        query = """
            SELECT
                id, user_id, friend_id, message, is_welcome, is_sent, is_read,
                created, sentiment, hour
            FROM messages where is_processed = True and id>= %s order by id
            LIMIT %s 
        """
        cursor.execute(query,(process_id, batch_size))
        batch_data = cursor.fetchall()
        print("fetch done")
        return batch_data
    except Exception as e:
        print("Fetch error", e)
    finally:
        if conn:
            cursor.close()
            conn.close()

def export_to_excel(batch_size = 1000, process_id =1 ):
    try:
        texts = fetch_data(process_id, batch_size)
        sheet_name = f"chat_analysis"
        rows = []
        filename = "sentiment_analysis.xlsx"
        
        for item in texts:

            #row = item["row"]

            rows.append({
                "id": item["id"],
                "user_id": item["user_id"],
                "friend_id": item["friend_id"],
                "message": item["message"],
                "is_read": item["is_read"],
                "is_sent": item["is_sent"],
                "is_welcome": item["is_welcome"],
                "created": item["created"],
                "sentiment": item["sentiment"],
                "hour" : item["hour"]
            })
        df = pd.DataFrame(rows)
        if not os.path.exists(filename):
            df.to_excel(filename, sheet_name=sheet_name, index=False)
        else:
            wb = load_workbook(filename)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name)
            
            for row in df.values.tolist():
                ws.append(row)
            wb.save(filename)
            wb.close()
            print(f"Appended data to existing sheet: '{sheet_name}'")

        print(f"Excel exported: {filename}")
    except Exception as e:
        print("Excel Error Occurred:", e)

#export_to_excel(10)
